from django.shortcuts import render , redirect, HttpResponse, get_object_or_404
from django.contrib.auth import authenticate , login , logout
from django.contrib import messages
from .models import Record , Comment , RecordResource, Folder, File
from .forms import AddRecordForms , CommentForm, ImportRecordDataForm, NewFolderForm
from django.views.generic.edit import CreateView, UpdateView
from django.views.generic import CreateView
from django.urls import reverse_lazy
from .resources import RecordResource
from django.contrib.auth.decorators import login_required
from tablib import Dataset
import pandas as pd
from rest_framework import generics
from rest_framework.parsers import MultiPartParser
from import_export.formats.base_formats import CSV , XLS , XLSX
from import_export import resources
from django.views.generic import View
from openpyxl import load_workbook
from .filters import RecordFilter
from django.utils import timezone
import os



def home(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, "Du er logget ind")
    records = Record.objects.all()
    return render(request, 'home.html', {'records': records})

def archived(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, "Du er logget ind")
    records = Record.objects.all()
    return render(request, 'archived.html', {'records': records})

def prospects(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, "Du er logget ind")
    records = Record.objects.all()
    return render(request, 'prospects.html', {'records': records})

class CustomerRecordView(View):
    def get(self, request, pk):
        if request.user.is_authenticated:
            customer_record = Record.objects.get(id=pk)
            folders = customer_record.folders.all()
            return render(request, "Record.html", {'customer_record':customer_record, 'folders':folders})
        else:
            messages.success(request, "Du skal være logget ind for at se siden")
            return redirect('home')

    def post(self, request, pk):
        if request.user.is_authenticated:
            customer_record = Record.objects.get(id=pk)
            messages.success(request, "Sagen er blevet gemt")
            return redirect('Record')
        else:
            messages.success(request, "Du skal være logget ind for at se siden")
            return redirect('home')

def login_user (request): 
	pass

def logout_user (request):
	logout(request)
	messages.success(request, "Du er nu logget ud")
	return redirect('home')


def delete_record (request, pk):
	if request.user.is_authenticated:
		delete_it = Record.objects.get(id=pk)
		delete_it.delete()
		messages.success(request, "Registrering er slettet")
		return redirect('home')

	else:
			messages.success(request, "Du skal være logget ind for at se siden")
			return redirect('home')


def add_record(request):
	form = AddRecordForms(request.POST or None)
	if request.user.is_authenticated:
		if request.method == "POST":
			if form.is_valid():
				add_record = form.save()
				messages.success(request, "Registrering er tilføjet...")
				return redirect('prospects')
		return render(request, 'add_record.html', {'form':form})
	else:
		messages.success(request, "Du skal være logget ind for at se siden")
		return redirect('home')

def update_record(request, pk):
    if request.user.is_authenticated:
        customer_record = Record.objects.get(id=pk)
        form = AddRecordForms(request.POST or None, instance=customer_record)
        if form.is_valid():
            form.save()
            messages.success(request, "Sagen er blevet opdateret")
            return redirect('Record', pk=customer_record.pk)
        return render(request, 'update_record.html', {'form':form})
    else:
        messages.success(request, "Du skal være logget ind for at se siden")
        return redirect('home')

def record_details(request, pk):
    record = Record.objects.get(id=pk)
    return render(request, 'Record.html', {'record': record})


class AddCommentView(CreateView):
    model = Comment
    form_class = CommentForm
    template_name = 'add_comment.html'

    def form_valid(self, form):
        form.instance.post_id = self.kwargs['pk']
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('Record', kwargs={'pk': self.kwargs['pk']})


class ImportRecordData(View):
    form_class = ImportRecordDataForm
    template_name = 'importrecorddata.html'

    def get(self, request):
        form = self.form_class(None)
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = self.form_class(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            # import record data logic goes here
            # ...
            return redirect('home')
        return render(request, self.template_name, {'form': form})


def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            created_at, BFE_Nummer, Adresse, Kommune, Region, Kontaktperson, Mail, Telefonnummer, m2, Kommuneplan, Lokalplan, Formaal = row
            customer_record.objects.create(created_at=created_at, BFE_Nummer=BFE_Nummer, Adresse=Adresse, Kommune=Kommune, Region=Region, Kontaktperson=Kontaktperson, Mail=Mail, Telefonnummer=Telefonnummer, m2=m2, Kommuneplan=Kommuneplan, Lokalplan=Lokalplan, Formål=Formål)

        return render(request, 'import_success.html')

    return render(request, 'import_form.html')

def create_new_folder(request):
    if request.method == 'POST':
        form = NewFolderForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponse('Mappe oprettet')
    else:
        createfolder = NewFolderForm
    return render(request, 'createnewfolder.html', {'createfolder':createfolder})

def open_folder(request, pk):
    folder = get_object_or_404(Folder, pk=pk)
    return render(request, 'openfolder.html', {'folder':folder})

def upload_file(request):
    uploaded_file = request.FILES.get('uploadfile')
    folder_id = request.POST.get('fid')
    folder = get_object_or_404(Folder, pk=folder_id)
    File.objects.create(folder=folder, files=uploaded_file)

    return redirect('open_folder', pk=folder_id)

class EditCommentView(UpdateView):
    model = Comment
    form_class = CommentForm
    template_name = 'edit_comment.html'

    def get_object(self, queryset=None):
        comment = get_object_or_404(Comment, id=self.kwargs['comment_id'], user=self.request.user)
        return comment

    def form_valid(self, form):
        form.instance.modified_on = timezone.now()
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('Record', kwargs={'pk': self.object.post.id})